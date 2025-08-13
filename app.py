import os
import sys
import logging
import time
from datetime import datetime, timedelta, timezone
from flask import Flask, request, jsonify, send_file, send_from_directory
from flask_cors import CORS
from pymongo import MongoClient
from gridfs import GridFS
from bson import ObjectId
from werkzeug.utils import secure_filename
import io
import openpyxl
import mimetypes
import pandas as pd
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from apscheduler.schedulers.background import BackgroundScheduler
# Workaround for numpy initialization issue in PyInstaller
try:
    import numpy as np
except RuntimeError as e:
    if "CPU dispatcher tracer already initialized" in str(e):
        import numpy.core.multiarray # Force reinitialization
        np = __import__('numpy')
    else:
        raise
scheduler = BackgroundScheduler()
# Initialize Flask app
def get_base_dir():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))
BASE_DIR = get_base_dir()
CONFIG_DIR = os.getenv('CONFIG_DIR', BASE_DIR)
os.makedirs(CONFIG_DIR, exist_ok=True)
CONFIG_FILE_PATH = os.path.join(CONFIG_DIR, 'config.json')
STATIC_FOLDER_PATH = os.path.join(BASE_DIR, 'dist')
BACKUP_DIR = os.getenv('BACKUP_DIR', os.path.join(BASE_DIR, 'backups'))
os.makedirs(BACKUP_DIR, exist_ok=True)
app = Flask(__name__, static_folder=STATIC_FOLDER_PATH, static_url_path='/')
CORS(app, resources={r"/api/*": {"origins": "http://localhost:5173"}}, supports_credentials=True)
# Configure logging
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)
# MongoDB connection
def connect_to_mongodb():
    try:
        client = MongoClient('mongodb://localhost:27017/')
        db = client['retail']
        return db
    except Exception as e:
        logger.error(f"Failed to connect to MongoDB: {e}")
        return None
db = connect_to_mongodb()
if db:
    fs = GridFS(db)
    products_collection = db['products']
    mobiles_collection = db['mobiles']
    accessories_collection = db['accessories']
    sales_collection = db['sales']
    customers_collection = db['customers']
    settings_collection = db['settings']
    print_collection = db['print']
    stock_additions_collection = db['stock_additions']
    email_collection = db['email']
    backup_collection = db['backup']
else:
    logger.error("MongoDB connection failed. Database operations will not work.")
    fs = None
    products_collection = None
    mobiles_collection = None
    accessories_collection = None
    sales_collection = None
    customers_collection = None
    settings_collection = None
    print_collection = None
    stock_additions_collection = None
    email_collection = None
    backup_collection = None
# Configure upload folder
UPLOAD_FOLDER = os.getenv('UPLOAD_FOLDER', 'Uploads')
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}
ALLOWED_EXCEL_EXTENSIONS = {'xlsx', 'xls'}
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS
def allowed_excel_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXCEL_EXTENSIONS
@app.route('/api/upload/images', methods=['POST'])
def upload_images():
    try:
        if not request.files:
            return jsonify({"error": "No images provided"}), 400
        uploaded_images = []
        for key in request.files:
            if key.startswith('images['):
                image_file = request.files[key]
                if image_file and image_file.filename != '' and allowed_file(image_file.filename):
                    filename = secure_filename(image_file.filename)
                    file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                    image_file.save(file_path)
                    logger.info(f"Image saved to {file_path}")
                    uploaded_images.append({
                        "filename": filename,
                        "path": file_path
                    })
                else:
                    logger.error(f"Invalid file: {image_file.filename}")
                    return jsonify({"error": f"Invalid file: {image_file.filename}"}), 400
        if not uploaded_images:
            return jsonify({"error": "No valid images uploaded"}), 400
        return jsonify({"message": "Images uploaded successfully", "uploadedImages": uploaded_images}), 201
    except Exception as e:
        logger.error(f"Error uploading images: {e}")
        return jsonify({"error": f"Failed to upload images: {str(e)}"}), 500
@app.route('/api/export/mobiles', methods=['GET'])
def export_mobiles_to_excel():
    try:
        headers = [
            "Name", "Price (₹)", "Stock Quantity", "Category", "Supplier", "Minimum Stock Level",
            "Barcode", "Model", "Type (Brand/General Model)", "Image ID", "Image Path", "Image"
        ]
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Mobile Products"
        ws.append(headers)
        # Fetch mobile products
        mobile_products = list(products_collection.find({"category": {"$regex": "^mobile$", "$options": "i"}}))
        for row_idx, product in enumerate(mobile_products, start=2):
            image_path = product.get('image_path')
            image_id = product.get('image_id')
            image_data = None
            # Retrieve image from GridFS or filesystem
            if image_id and ObjectId.is_valid(image_id) and fs.exists(ObjectId(image_id)):
                image = fs.get(ObjectId(image_id))
                image_data = io.BytesIO(image.read())
                logger.info(f"Image retrieved from GridFS: {image_id}")
            elif image_path:
                full_path = os.path.join(app.config['UPLOAD_FOLDER'], image_path)
                if os.path.exists(full_path):
                    with open(full_path, 'rb') as f:
                        image_data = io.BytesIO(f.read())
                        logger.info(f"Image retrieved from filesystem: {full_path}")
                else:
                    logger.warning(f"Image not found at {full_path}")
            # Populate row data
            ws.cell(row=row_idx, column=1).value = product.get('name', '')
            ws.cell(row=row_idx, column=2).value = product.get('price', 0)
            ws.cell(row=row_idx, column=3).value = product.get('stock', 0)
            ws.cell(row=row_idx, column=4).value = product.get('category', '')
            ws.cell(row=row_idx, column=5).value = product.get('supplier', '')
            ws.cell(row=row_idx, column=6).value = product.get('minStock', 5)
            ws.cell(row=row_idx, column=7).value = product.get('barcode', '')
            ws.cell(row=row_idx, column=8).value = product.get('model', '')
            ws.cell(row=row_idx, column=9).value = product.get('type', '')
            ws.cell(row=row_idx, column=10).value = str(image_id) if image_id else ''
            ws.cell(row=row_idx, column=11).value = image_path if image_path else ''
            # Embed image
            if image_data:
                img = openpyxl.drawing.image.Image(image_data)
                img.width = 100
                img.height = 100
                cell = ws.cell(row=row_idx, column=12)
                ws.add_image(img, f"{cell.column_letter}{cell.row}")
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width
        # Adjust row heights for images
        for row in range(2, len(mobile_products) + 2):
            ws.row_dimensions[row].height = 80
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='mobile_products.xlsx'
        )
    except Exception as e:
        logger.error(f"Error exporting mobiles to Excel: {e}")
        return jsonify({"error": "Failed to export mobile products"}), 500
@app.route('/api/export/accessories', methods=['GET'])
def export_accessories_to_excel():
    try:
        headers = [
            "Name", "Price (₹)", "Stock Quantity", "Category", "Supplier", "Minimum Stock Level",
            "Barcode", "Type", "Accessory Type", "Image ID", "Image Path", "Image"
        ]
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Accessories Products"
        ws.append(headers)
        # Fetch accessory products
        accessory_products = list(products_collection.find({"category": {"$regex": "^accessories$", "$options": "i"}}))
        for row_idx, product in enumerate(accessory_products, start=2):
            image_path = product.get('image_path')
            image_id = product.get('image_id')
            image_data = None
            # Retrieve image from GridFS or filesystem
            if image_id and ObjectId.is_valid(image_id) and fs.exists(ObjectId(image_id)):
                image = fs.get(ObjectId(image_id))
                image_data = io.BytesIO(image.read())
                logger.info(f"Image retrieved from GridFS: {image_id}")
            elif image_path:
                full_path = os.path.join(app.config['UPLOAD_FOLDER'], image_path)
                if os.path.exists(full_path):
                    with open(full_path, 'rb') as f:
                        image_data = io.BytesIO(f.read())
                        logger.info(f"Image retrieved from filesystem: {full_path}")
                else:
                    logger.warning(f"Image not found at {full_path}")
            # Populate row data
            ws.cell(row=row_idx, column=1).value = product.get('name', '')
            ws.cell(row=row_idx, column=2).value = product.get('price', 0)
            ws.cell(row=row_idx, column=3).value = product.get('stock', 0)
            ws.cell(row=row_idx, column=4).value = product.get('category', '')
            ws.cell(row=row_idx, column=5).value = product.get('supplier', '')
            ws.cell(row=row_idx, column=6).value = product.get('minStock', 5)
            ws.cell(row=row_idx, column=7).value = product.get('barcode', '')
            ws.cell(row=row_idx, column=8).value = product.get('type', '')
            ws.cell(row=row_idx, column=9).value = product.get('accessoryType', '')
            ws.cell(row=row_idx, column=10).value = str(image_id) if image_id else ''
            ws.cell(row=row_idx, column=11).value = image_path if image_path else ''
            # Embed image
            if image_data:
                img = openpyxl.drawing.image.Image(image_data)
                img.width = 100
                img.height = 100
                cell = ws.cell(row=row_idx, column=12)
                ws.add_image(img, f"{cell.column_letter}{cell.row}")
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width
        # Adjust row heights for images
        for row in range(2, len(accessory_products) + 2):
            ws.row_dimensions[row].height = 80
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='accessory_products.xlsx'
        )
    except Exception as e:
        logger.error(f"Error exporting accessories to Excel: {e}")
        return jsonify({"error": "Failed to export accessory products"}), 500
# Serve React app
@app.route('/', defaults={'path': ''})
@app.route('/<path:path>')
def serve_react_app(path):
    if path != "" and os.path.exists(os.path.join(app.static_folder, path)):
        return send_from_directory(app.static_folder, path)
    return send_from_directory(app.static_folder, 'index.html')
# Print Settings Routes
@app.route('/api/print', methods=['GET'])
def get_print_settings():
    try:
        print_settings = print_collection.find_one({"_id": "print_settings"})
        if print_settings:
            return jsonify({
                "shopName": print_settings.get("shopName", "Your Shop Name"),
                "address": print_settings.get("address", "123 Shop Street, City, Country"),
                "gstin": print_settings.get("gstin", "12ABCDE1234F1Z5")
            })
        else:
            return jsonify({
                "shopName": "Your Shop Name",
                "address": "123 Shop Street, City, Country",
                "gstin": "12ABCDE1234F1Z5"
            })
    except Exception as e:
        logger.error(f"Error fetching print settings: {e}")
        return jsonify({"error": str(e)}), 500
@app.route('/api/print', methods=['POST'])
def save_print_settings():
    try:
        data = request.get_json()
        shop_name = data.get('shopName')
        address = data.get('address')
        gstin = data.get('gstin')
        if not shop_name or not address or not gstin:
            return jsonify({"error": "All fields are required"}), 400
        print_collection.update_one(
            {"_id": "print_settings"},
            {"$set": {
                "shopName": shop_name,
                "address": address,
                "gstin": gstin
            }},
            upsert=True
        )
        return jsonify({
            "shopName": shop_name,
            "address": address,
            "gstin": gstin,
            "message": "Print settings saved successfully"
        })
    except Exception as e:
        logger.error(f"Error saving print settings: {e}")
        return jsonify({"error": str(e)}), 500
# Settings Routes
@app.route('/api/settings', methods=['GET'])
def get_settings():
    try:
        settings = settings_collection.find_one({"key": "gst_settings"})
        if settings:
            return jsonify({
                "gstPercentage": settings.get("gstPercentage", 18),
                "enableGst": settings.get("enableGst", True)
            })
        return jsonify({"gstPercentage": 18, "enableGst": True})
    except Exception as e:
        logger.error(f"Error fetching settings: {e}")
        return jsonify({"error": "Failed to fetch settings"}), 500
@app.route('/api/settings', methods=['POST'])
def save_settings():
    try:
        data = request.json
        gst_percentage = float(data.get('gstPercentage', 18))
        enable_gst = data.get('enableGst', True)
        if gst_percentage < 0:
            return jsonify({"error": "GST percentage cannot be negative"}), 400
        settings_collection.update_one(
            {"key": "gst_settings"},
            {"$set": {"gstPercentage": gst_percentage, "enableGst": enable_gst}},
            upsert=True
        )
        return jsonify({
            "message": "Settings saved successfully",
            "gstPercentage": gst_percentage,
            "enableGst": enable_gst
        }), 200
    except Exception as e:
        logger.error(f"Error saving settings: {e}")
        return jsonify({"error": "Failed to save settings"}), 500
# Image Serving Route
@app.route('/api/images/<image_id>')
def serve_image(image_id):
    try:
        logger.info(f"Attempting to serve image with ID: {image_id}")
        if not ObjectId.is_valid(image_id):
            logger.error(f"Invalid image ID format: {image_id}")
            return jsonify({"error": "Invalid image ID"}), 400
        if not fs.exists(ObjectId(image_id)):
            logger.error(f"Image not found in GridFS: {image_id}")
            return jsonify({"error": "Image not found"}), 404
        image = fs.get(ObjectId(image_id))
        content_type = image.content_type
        if not content_type and image.filename:
            content_type, _ = mimetypes.guess_type(image.filename)
        if not content_type:
            content_type = 'application/octet-stream'
        logger.info(f"Serving image {image_id} from GridFS with content_type: {content_type}")
        return send_file(io.BytesIO(image.read()), mimetype=content_type)
    except Exception as e:
        logger.error(f"Error serving image {image_id}: {e}")
        return jsonify({"error": "Image not found"}), 404
# Products Routes
@app.route('/api/products', methods=['GET'])
def get_products():
    try:
        products = list(products_collection.find({}))
        for product in products:
            product['_id'] = str(product['_id'])
            product['image'] = None
            image_path = product.get('image_path')
            if image_path:
                full_path = os.path.join(app.config['UPLOAD_FOLDER'], image_path)
                if os.path.exists(full_path):
                    product['image'] = f"/Uploads/{image_path}"
        return jsonify(products)
    except Exception as e:
        logger.error(f"Error fetching products: {e}")
        return jsonify({"error": "An unexpected error occurred"}), 500
@app.route('/api/products', methods=['POST'])
def add_product():
    try:
        data = request.form.to_dict()
        name = data.get('name')
        price = float(data.get('price', 0))
        stock = int(data.get('stock', 0))
        category = data.get('category')
        supplier = data.get('supplier', '')
        min_stock = int(data.get('minStock', 5))
        barcode = data.get('barcode', f"BC{int(time.time())}")
        model = data.get('model', '')
        accessory_type = data.get('accessoryType', '')
        new_mobile = data.get('newMobile', '')
        new_accessory_name = data.get('newAccessoryName', '')
        new_accessory_model = data.get('newAccessoryModel', '')
        product_type = data.get('type', '')
        image_path = data.get('image_path', '')
        if not name or not category:
            return jsonify({"error": "Name and category are required"}), 400
        if 'image' in request.files:
            image_file = request.files['image']
            if image_file and image_file.filename != '' and allowed_file(image_file.filename):
                filename = secure_filename(image_file.filename)
                file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                image_file.save(file_path)
                image_path = filename
        if category.lower() == "mobile" and new_mobile:
            existing_mobile = mobiles_collection.find_one({"name": new_mobile})
            if not existing_mobile:
                mobiles_collection.insert_one({"name": new_mobile, "category": "Mobile"})
            model = new_mobile
        if category.lower() == "accessories" and new_accessory_name and new_accessory_model:
            accessory_full_name = f"{new_accessory_model} - {new_accessory_name}"
            existing_accessory = accessories_collection.find_one({"accessoryName": new_accessory_name, "accessoryModel": new_accessory_model})
            if not existing_accessory:
                accessories_collection.insert_one({
                    "accessoryName": new_accessory_name,
                    "accessoryModel": new_accessory_model,
                    "category": "Accessories",
                    "type": product_type
                })
            accessory_type = accessory_full_name
        product_data = {
            "name": name,
            "price": price,
            "stock": stock,
            "category": category,
            "supplier": supplier,
            "minStock": min_stock,
            "barcode": barcode,
            "image_path": image_path or None,
            "model": model if category.lower() == "mobile" else "",
            "accessoryType": accessory_type if category.lower() == "accessories" else "",
            "type": product_type
        }
        result = products_collection.insert_one(product_data)
        if stock > 0:
            stock_additions_collection.insert_one({
                "product_id": str(result.inserted_id),
                "quantity": stock,
                "date": datetime.now(timezone.utc).isoformat()
            })
        return jsonify({"message": "Product added successfully!", "id": str(result.inserted_id)}), 201
    except Exception as e:
        logger.error(f"Error adding product: {e}")
        return jsonify({"error": "An unexpected error occurred"}), 500
@app.route('/api/products/<id>', methods=['PUT'])
def update_product(id):
    try:
        if not ObjectId.is_valid(id):
            return jsonify({"error": "Invalid product ID"}), 400
        data = request.form.to_dict()
        update_fields = {}
        if 'name' in data: update_fields['name'] = data['name'] or ''
        if 'price' in data: update_fields['price'] = float(data['price']) if data['price'] else 0
        if 'stock' in data: update_fields['stock'] = int(data['stock']) if data['stock'] else 0
        if 'category' in data: update_fields['category'] = data['category'] or ''
        if 'supplier' in data: update_fields['supplier'] = data['supplier'] or ''
        if 'minStock' in data: update_fields['minStock'] = int(data['minStock']) if data['minStock'] else 5
        if 'barcode' in data: update_fields['barcode'] = data['barcode'] or f"BC{int(time.time())}"
        if 'type' in data: update_fields['type'] = data['type'] or ''
        if 'image' in request.files:
            image_file = request.files['image']
            if image_file and image_file.filename != '' and allowed_file(image_file.filename):
                existing_product = products_collection.find_one({"_id": ObjectId(id)})
                if existing_product and 'image_id' in existing_product and existing_product['image_id'] and ObjectId.is_valid(existing_product['image_id']):
                    fs.delete(ObjectId(existing_product['image_id']))
                filename = secure_filename(image_file.filename)
                content_type, _ = mimetypes.guess_type(filename)
                if content_type is None:
                    content_type = 'application/octet-stream'
                new_image_id = fs.put(image_file, filename=filename, content_type=content_type)
                update_fields['image_id'] = str(new_image_id)
                update_fields['image_path'] = filename
        elif 'image_path' in data and data['image_path'] == '':
            existing_product = products_collection.find_one({"_id": ObjectId(id)})
            if existing_product and 'image_id' in existing_product and existing_product['image_id'] and ObjectId.is_valid(existing_product['image_id']):
                fs.delete(ObjectId(existing_product['image_id']))
            update_fields['image_id'] = None
            update_fields['image_path'] = None
        current_product = products_collection.find_one({"_id": ObjectId(id)})
        if not current_product:
            return jsonify({"error": "Product not found"}), 404
        current_category = update_fields.get('category', current_product['category'])
        if current_category.lower() == "mobile":
            new_mobile = data.get('newMobile', '')
            if new_mobile:
                existing_mobile = mobiles_collection.find_one({"name": new_mobile})
                if not existing_mobile:
                    mobiles_collection.insert_one({"name": new_mobile, "category": "Mobile"})
                update_fields['model'] = new_mobile
            elif 'model' in data: update_fields['model'] = data['model'] or ''
            update_fields['accessoryType'] = ""
        elif current_category.lower() == "accessories":
            new_accessory_name = data.get('newAccessoryName', '')
            new_accessory_model = data.get('newAccessoryModel', '')
            if new_accessory_name and new_accessory_model:
                accessory_full_name = f"{new_accessory_model} - {new_accessory_name}"
                existing_accessory = accessories_collection.find_one({"accessoryName": new_accessory_name, "accessoryModel": new_accessory_model})
                if not existing_accessory:
                    accessories_collection.insert_one({"accessoryName": new_accessory_name, "accessoryModel": new_accessory_model, "category": "Accessories"})
                update_fields['accessoryType'] = accessory_full_name
            elif 'accessoryType' in data: update_fields['accessoryType'] = data['accessoryType'] or ''
            update_fields['model'] = ""
        else:
            update_fields['model'] = ""
            update_fields['accessoryType'] = ""
        if 'stock' in update_fields and update_fields['stock'] != current_product['stock']:
            quantity_change = update_fields['stock'] - current_product['stock']
            if quantity_change > 0:
                stock_additions_collection.insert_one({
                    "product_id": id,
                    "quantity": quantity_change,
                    "date": datetime.utcnow().isoformat()
                })
        result = products_collection.update_one({"_id": ObjectId(id)}, {"$set": update_fields})
        if result.matched_count == 0:
            return jsonify({"error": "Product not found"}), 404
        return jsonify({"message": "Product updated successfully!"})
    except Exception as e:
        logger.error(f"Error updating product {id}: {e}")
        return jsonify({"error": "An unexpected error occurred"}), 500
@app.route('/api/products/<id>', methods=['DELETE'])
def delete_product(id):
    try:
        product = products_collection.find_one({"_id": ObjectId(id)})
        if not product:
            return jsonify({"error": "Product not found"}), 404
        if 'image_id' in product and product['image_id'] and ObjectId.is_valid(product['image_id']):
            fs.delete(ObjectId(product['image_id']))
        products_collection.delete_one({"_id": ObjectId(id)})
        return jsonify({"message": "Product deleted successfully!"})
    except Exception as e:
        logger.error(f"Error deleting product {id}: {e}")
        return jsonify({"error": "An unexpected error occurred"}), 500
@app.route('/api/products/all', methods=['DELETE'])
def delete_all_products():
    try:
        products = list(products_collection.find())
        for product in products:
            if 'image_id' in product and product['image_id'] and ObjectId.is_valid(product['image_id']):
                fs.delete(ObjectId(product['image_id']))
        result = products_collection.delete_many({})
        return jsonify({"message": f"Deleted {result.deleted_count} products successfully!"})
    except Exception as e:
        logger.error(f"Error deleting all products: {e}")
        return jsonify({"error": "An unexpected error occurred"}), 500
@app.route('/api/products/mobile', methods=['DELETE'])
def delete_mobile_products():
    try:
        products = list(products_collection.find({"category": {"$regex": "^mobile$", "$options": "i"}}))
        for product in products:
            if 'image_id' in product and product['image_id'] and ObjectId.is_valid(product['image_id']):
                fs.delete(ObjectId(product['image_id']))
        result = products_collection.delete_many({"category": {"$regex": "^mobile$", "$options": "i"}})
        return jsonify({"message": f"Deleted {result.deleted_count} mobile products successfully!"})
    except Exception as e:
        logger.error(f"Error deleting mobile products: {e}")
        return jsonify({"error": "An unexpected error occurred"}), 500
@app.route('/api/products/accessories', methods=['DELETE'])
def delete_accessories_products():
    try:
        products = list(products_collection.find({"category": {"$regex": "^accessories$", "$options": "i"}}))
        for product in products:
            if 'image_id' in product and product['image_id'] and ObjectId.is_valid(product['image_id']):
                fs.delete(ObjectId(product['image_id']))
        result = products_collection.delete_many({"category": {"$regex": "^accessories$", "$options": "i"}})
        return jsonify({"message": f"Deleted {result.deleted_count} accessories products successfully!"})
    except Exception as e:
        logger.error(f"Error deleting accessories products: {e}")
        return jsonify({"error": "An unexpected error occurred"}), 500
@app.route('/api/products/category/<category>', methods=['DELETE'])
def delete_products_by_category(category):
    try:
        products = list(products_collection.find({"category": {"$regex": f"^{category}$", "$options": "i"}}))
        for product in products:
            if 'image_id' in product and product['image_id'] and ObjectId.is_valid(product['image_id']):
                fs.delete(ObjectId(product['image_id']))
        result = products_collection.delete_many({"category": {"$regex": f"^{category}$", "$options": "i"}})
        return jsonify({"message": f"Deleted {result.deleted_count} products in category {category} successfully!"})
    except Exception as e:
        logger.error(f"Error deleting products in category {category}: {e}")
        return jsonify({"error": "An unexpected error occurred"}), 500
@app.route('/api/products/<id>/stock', methods=['PUT'])
def update_product_stock(id):
    try:
        data = request.get_json()
        new_stock = int(data.get('stock', 0))
        current_product = products_collection.find_one({"_id": ObjectId(id)})
        if not current_product:
            return jsonify({"error": "Product not found"}), 404
        quantity_change = new_stock - current_product['stock']
        if quantity_change > 0:
            stock_additions_collection.insert_one({
                "product_id": id,
                "quantity": quantity_change,
                "date": datetime.utcnow().isoformat()
            })
        result = products_collection.update_one({"_id": ObjectId(id)}, {"$set": {"stock": new_stock}})
        if result.matched_count == 0:
            return jsonify({"error": "Product not found"}), 404
        return jsonify({"message": "Stock updated successfully!"})
    except Exception as e:
        logger.error(f"Error updating stock for product {id}: {e}")
        return jsonify({"error": "An unexpected error occurred"}), 500
# Mobiles Routes
@app.route('/api/mobiles', methods=['GET'])
def get_mobiles():
    try:
        mobiles = list(mobiles_collection.find({}))
        for mobile in mobiles:
            mobile['_id'] = str(mobile['_id'])
        return jsonify(mobiles)
    except Exception as e:
        logger.error(f"Error fetching mobiles: {e}")
        return jsonify({"error": "An unexpected error occurred"}), 500
@app.route('/api/mobiles', methods=['POST'])
def add_mobile():
    try:
        data = request.form.to_dict()
        name = data.get('name')
        if not name:
            return jsonify({"error": "Mobile name is required"}), 400
        mobiles_collection.insert_one({"name": name, "category": "Mobile"})
        return jsonify({"message": "Mobile type added successfully!"}), 201
    except Exception as e:
        logger.error(f"Error adding mobile: {e}")
        return jsonify({"error": "An unexpected error occurred"}), 500
@app.route('/api/mobiles/<id>', methods=['PUT'])
def update_mobile(id):
    try:
        data = request.form.to_dict()
        name = data.get('name')
        if not name:
            return jsonify({"error": "Mobile name is required"}), 400
        result = mobiles_collection.update_one({"_id": ObjectId(id)}, {"$set": {"name": name}})
        if result.matched_count == 0:
            return jsonify({"error": "Mobile not found"}), 404
        return jsonify({"message": "Mobile type updated successfully!"})
    except Exception as e:
        logger.error(f"Error updating mobile {id}: {e}")
        return jsonify({"error": "An unexpected error occurred"}), 500
@app.route('/api/mobiles/<id>', methods=['DELETE'])
def delete_mobile(id):
    try:
        result = mobiles_collection.delete_one({"_id": ObjectId(id)})
        if result.deleted_count == 0:
            return jsonify({"error": "Mobile not found"}), 404
        return jsonify({"message": "Mobile type deleted successfully!"})
    except Exception as e:
        logger.error(f"Error deleting mobile {id}: {e}")
        return jsonify({"error": "An unexpected error occurred"}), 500
# Accessories Routes
@app.route('/api/accessories', methods=['GET'])
def get_accessories():
    try:
        accessories = list(accessories_collection.find({}))
        for accessory in accessories:
            accessory['_id'] = str(accessory['_id'])
        return jsonify(accessories)
    except Exception as e:
        logger.error(f"Error fetching accessories: {e}")
        return jsonify({"error": "An unexpected error occurred"}), 500
@app.route('/api/accessories', methods=['POST'])
def add_accessory():
    try:
        data = request.form.to_dict()
        accessory_name = data.get('accessoryName')
        accessory_model = data.get('accessoryModel')
        category = data.get('category', 'Accessories')
        accessory_type_field = data.get('type', '')
        if not all([accessory_name, accessory_model]):
            return jsonify({"error": "Accessory name and model are required"}), 400
        existing_accessory = accessories_collection.find_one({"accessoryName": accessory_name, "accessoryModel": accessory_model})
        if existing_accessory:
            return jsonify({"error": "Accessory type with this name and model already exists"}), 400
        accessory_data = {
            'accessoryName': accessory_name,
            'accessoryModel': accessory_model,
            'category': category,
            'type': accessory_type_field
        }
        result = accessories_collection.insert_one(accessory_data)
        return jsonify({"message": "Accessory type added successfully!", "id": str(result.inserted_id)}), 201
    except Exception as e:
        logger.error(f"Error adding accessory: {e}")
        return jsonify({"error": "An unexpected error occurred"}), 500
@app.route('/api/accessories/<id>', methods=['PUT'])
def update_accessory(id):
    try:
        data = request.form.to_dict()
        accessory_name = data.get('accessoryName')
        accessory_model = data.get('accessoryModel')
        category = data.get('category', 'Accessories')
        accessory_type_field = data.get('type', '')
        if not all([accessory_name, accessory_model]):
            return jsonify({"error": "Accessory name and model are required"}), 400
        existing_accessory = accessories_collection.find_one({
            "accessoryName": accessory_name,
            "accessoryModel": accessory_model,
            "_id": {"$ne": ObjectId(id)}
        })
        if existing_accessory:
            return jsonify({"error": "Accessory type with this name and model already exists"}), 400
        update_data = {
            'accessoryName': accessory_name,
            'accessoryModel': accessory_model,
            'category': category,
            'type': accessory_type_field
        }
        result = accessories_collection.update_one({"_id": ObjectId(id)}, {"$set": update_data})
        if result.matched_count == 0:
            return jsonify({"error": "Accessory not found"}), 404
        return jsonify({"message": "Accessory type updated successfully!"})
    except Exception as e:
        logger.error(f"Error updating accessory {id}: {e}")
        return jsonify({"error": "An unexpected error occurred"}), 500
@app.route('/api/accessories/<id>', methods=['DELETE'])
def delete_accessory(id):
    try:
        result = accessories_collection.delete_one({"_id": ObjectId(id)})
        if result.deleted_count == 0:
            return jsonify({"error": "Accessory not found"}), 404
        return jsonify({"message": "Accessory type deleted successfully!"})
    except Exception as e:
        logger.error(f"Error deleting accessory {id}: {e}")
        return jsonify({"error": "An unexpected error occurred"}), 500
# Image Retrieval Route
@app.route('/Uploads/<filename>')
def uploaded_file(filename):
    try:
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        if not os.path.exists(file_path):
            return jsonify({"error": "File not found"}), 404
        mimetype, _ = mimetypes.guess_type(filename)
        if not mimetype:
            mimetype = 'application/octet-stream'
        return send_file(file_path, mimetype=mimetype)
    except Exception as e:
        logger.error(f"Error serving file {filename}: {e}")
        return jsonify({"error": "Failed to retrieve image"}), 500
# Sales Route
@app.route('/api/sales', methods=['POST'])
def process_sale():
    try:
        data = request.json
        required_fields = ['customer', 'items', 'subtotal', 'tax', 'total', 'paymentMethod', 'timestamp', 'invoiceId', 'gstPercentage']
        if not all(field in data for field in required_fields):
            return jsonify({"error": "Missing required fields for sale"}), 400
        customer = data['customer']
        items = data['items']
        gst_percentage = data['gstPercentage']
        for item in items:
            product = products_collection.find_one({"_id": ObjectId(item['id'])})
            if not product:
                return jsonify({"error": f"Product {item['name']} not found"}), 404
            if product['stock'] < item['quantity']:
                return jsonify({"error": f"Insufficient stock for {item['name']}. Available: {product['stock']}"}), 400
        for item in items:
            products_collection.update_one(
                {"_id": ObjectId(item['id'])},
                {"$inc": {"stock": -item['quantity']}}
            )
        customer_id = customer.get('id')
        if customer_id and ObjectId.is_valid(customer_id):
            customer_doc = customers_collection.find_one({"_id": ObjectId(customer_id)})
            if customer_doc:
                customers_collection.update_one(
                    {"_id": ObjectId(customer_id)},
                    {
                        "$inc": {"purchases": 1, "totalPurchases": data['total']},
                        "$set": {"lastPurchase": data['timestamp'][:10], "name": customer['name'], "phone": customer['phone']}
                    }
                )
            else:
                customer_id = None
        if not customer_id:
            new_customer = {
                'name': customer['name'],
                'phone': customer['phone'],
                'email': '', 'address': '', 'city': '', 'pincode': '', 'dateOfBirth': '',
                'purchases': 1, 'totalPurchases': data['total'], 'totalPurchases': data['total'], 'lastPurchase': data['timestamp'][:10], 'posBalance': 0
            }
            result = customers_collection.insert_one(new_customer)
            customer_id = str(result.inserted_id)
        sale_data = {
            'customer': {'id': customer_id, 'name': customer['name'], 'phone': customer['phone']},
            'items': items,
            'subtotal': data['subtotal'],
            'tax': data['tax'],
            'total': data['total'],
            'paymentMethod': data['paymentMethod'],
            'timestamp': data['timestamp'],
            'invoiceId': data['invoiceId'],
            'status': 'completed',
            'gstPercentage': gst_percentage
        }
        result = sales_collection.insert_one(sale_data)
        return jsonify({"message": "Sale processed successfully", "saleId": str(result.inserted_id)}), 201
    except Exception as e:
        logger.error(f"Error in process_sale: {e}")
        return jsonify({"error": "Failed to process sale"}), 500
@app.route('/api/sales', methods=['GET'])
def get_sales():
    try:
        sales = list(sales_collection.find())
        for sale in sales:
            sale['_id'] = str(sale['_id'])
            sale['timestamp'] = sale['timestamp'][:10]
        return jsonify(sales)
    except Exception as e:
        logger.error(f"Error in get_sales: {e}")
        return jsonify({"error": "Failed to retrieve sales records"}), 500
# Dashboard Routes
@app.route('/api/dashboard/stats', methods=['GET'])
def get_dashboard_stats():
    try:
        total_sales = sum(sale.get('total', 0) for sale in sales_collection.find())
        unique_customers_count = customers_collection.count_documents({})
        low_stock_count = products_collection.count_documents({"$expr": {"$lte": ["$stock", "$minStock"]}})
        today = datetime.utcnow().date().isoformat()
        today_sales = sum(sale.get('total', 0) for sale in sales_collection.find({"timestamp": {"$regex": f"^{today}"}}))
        return jsonify({
            "todaySales": today_sales,
            "totalCustomers": unique_customers_count,
            "lowStockItems": low_stock_count,
            "totalRevenue": total_sales
        })
    except Exception as e:
        logger.error(f"Error in get_dashboard_stats: {e}")
        return jsonify({"error": "Failed to retrieve dashboard statistics"}), 500
@app.route('/api/dashboard/recent-sales', methods=['GET'])
def get_recent_sales():
    try:
        recent_sales = list(sales_collection.find().sort("timestamp", -1).limit(4))
        now = datetime.utcnow()
        processed_sales = []
        for sale in recent_sales:
            try:
                if not all(key in sale for key in ['timestamp', 'items', 'customer', 'total']):
                    logger.warning(f"Skipping sale {sale.get('_id')} due to missing fields")
                    continue
                sale_time = datetime.fromisoformat(sale['timestamp'].replace('Z', '+00:00'))
                time_diff = now - sale_time
                if time_diff < timedelta(minutes=60):
                    time_str = f"{int(time_diff.total_seconds() // 60)} mins ago"
                elif time_diff < timedelta(hours=24):
                    time_str = f"{int(time_diff.total_seconds() // 3600)} hours ago"
                else:
                    time_str = f"{int(time_diff.total_seconds() // 86400)} days ago"
                processed_sale = {
                    '_id': str(sale['_id']),
                    'time': time_str,
                    'product': sale['items'][0]['name'] if sale['items'] else 'Unknown',
                    'amount': f"₹{sale['total']:.0f}",
                    'customer': sale['customer']['name'] if 'name' in sale['customer'] else 'Unknown',
                    'status': sale.get('status', 'completed')
                }
                processed_sales.append(processed_sale)
            except Exception as e:
                logger.error(f"Error processing sale {sale.get('_id')}: {e}")
        return jsonify(processed_sales)
    except Exception as e:
        logger.error(f"Error in get_recent_sales: {e}")
        return jsonify({"error": "Failed to retrieve recent sales"}), 500
@app.route('/api/dashboard/low-stock', methods=['GET'])
def get_low_stock():
    try:
        low_stock_items = list(products_collection.find(
            {"$expr": {"$lte": ["$stock", "$minStock"]}},
            {"name": 1, "stock": 1, "minStock": 1, "_id": 1, "image_id": 1, "image_path": 1}
        ).limit(4))
        for item in low_stock_items:
            item['_id'] = str(item['_id'])
            image_id_str = item.get('image_id')
            image_path_str = item.get('image_path')
            item['image'] = None
            if image_id_str and ObjectId.is_valid(image_id_str):
                try:
                    if fs.exists(ObjectId(image_id_str)):
                        item['image'] = f"/api/images/{image_id_str}"
                except Exception as e:
                    logger.warning(f"Error accessing GridFS for low stock item {item['_id']} (image_id: {image_id_str}): {e}")
            if item['image'] is None and image_path_str:
                if not image_path_str.startswith('/') and not image_path_str.startswith('http'):
                    static_file_full_path = os.path.join(app.config['UPLOAD_FOLDER'], image_path_str)
                    if os.path.exists(static_file_full_path):
                        item['image'] = f"/Uploads/{image_path_str}"
                elif image_path_str.startswith('/api/images/'):
                    item['image'] = image_path_str
                elif image_path_str.startswith('http'):
                    item['image'] = image_path_str
        return jsonify(low_stock_items)
    except Exception as e:
        logger.error(f"Error in get_low_stock: {e}")
        return jsonify({"error": "Failed to retrieve low stock items"}), 500
@app.route('/api/dashboard/restock', methods=['POST'])
def restock_low_stock():
    try:
        data = request.get_json() or {}
        restock_amount = int(data.get('restockAmount', 10))
        low_stock_items = list(products_collection.find({"$expr": {"$lte": ["$stock", "$minStock"]}}))
   
        if not low_stock_items:
            return jsonify({"message": "No low stock items to restock", "updatedItems": []}), 200
        updated_items = []
        for item in low_stock_items:
            new_stock = item['minStock'] + restock_amount
            result = products_collection.update_one(
                {"_id": item['_id']},
                {"$set": {"stock": new_stock}}
            )
            if result.matched_count > 0:
                stock_additions_collection.insert_one({
                    "product_id": str(item['_id']),
                    "quantity": restock_amount,
                    "date": datetime.utcnow().isoformat()
                })
                updated_items.append({
                    "_id": str(item['_id']),
                    "name": item['name'],
                    "oldStock": item['stock'],
                    "newStock": new_stock,
                    "minStock": item['minStock']
                })
        return jsonify({
            "message": f"Successfully restocked {len(updated_items)} items",
            "updatedItems": updated_items
        }), 200
    except Exception as e:
        logger.error(f"Error in restock_low_stock: {e}")
        return jsonify({"error": "Failed to restock items"}), 500
@app.route('/api/dashboard/restock-manual', methods=['POST'])
def restock_manual():
    try:
        data = request.get_json() or {}
        items = data.get('items', [])
   
        if not items:
            return jsonify({"error": "No items provided for restocking"}), 400
        updated_items = []
        errors = []
        for item in items:
            item_id = item.get('itemId')
            quantity = item.get('quantity')
            min_stock = item.get('minStock')
            if not ObjectId.is_valid(item_id):
                errors.append(f"Invalid item ID: {item_id}")
                continue
            if not isinstance(quantity, int) or quantity <= 0:
                errors.append(f"Invalid quantity for item ID {item_id}")
                continue
            update_fields = {"stock": quantity}
            if min_stock is not None:
                if not isinstance(min_stock, int) or min_stock < 1:
                    errors.append(f"Invalid minStock for item ID {item_id}")
                    continue
                update_fields['minStock'] = min_stock
            product = products_collection.find_one({"_id": ObjectId(item_id)})
            if not product:
                errors.append(f"Product not found for item ID {item_id}")
                continue
            quantity_change = quantity - product['stock']
            if quantity_change > 0:
                stock_additions_collection.insert_one({
                    "product_id": item_id,
                    "quantity": quantity_change,
                    "date": datetime.utcnow().isoformat()
                })
            result = products_collection.update_one(
                {"_id": ObjectId(item_id)},
                {"$set": update_fields}
            )
            if result.matched_count > 0:
                updated_items.append({
                    "_id": str(item_id),
                    "name": product['name'],
                    "oldStock": product['stock'],
                    "newStock": quantity,
                    "minStock": min_stock if min_stock is not None else product['minStock']
                })
        if errors:
            return jsonify({
                "message": f"Restocked {len(updated_items)} items with {len(errors)} errors",
                "updatedItems": updated_items,
                "errors": errors
            }), 200
        return jsonify({
            "message": f"Successfully restocked {len(updated_items)} items",
            "updatedItems": updated_items
        }), 200
    except Exception as e:
        logger.error(f"Error in restock_manual: {e}")
        return jsonify({"error": "Failed to restock items"}), 500
# Customers Routes
@app.route('/api/customers', methods=['GET'])
def get_customers():
    try:
        customers = list(customers_collection.find())
        for customer in customers:
            customer['_id'] = str(customer['_id'])
        return jsonify(customers)
    except Exception as e:
        logger.error(f"Error in get_customers: {e}")
        return jsonify({"error": "Failed to retrieve customers"}), 500
@app.route('/api/customers', methods=['POST'])
def add_customer():
    try:
        data = request.json
        name = data.get('name')
        phone = data.get('phone')
        if not all([name, phone]):
            return jsonify({"error": "Name and phone number are required"}), 400
        customer_data = {
            'name': name,
            'phone': phone,
            'email': data.get('email', ''),
            'address': data.get('address', ''),
            'city': data.get('city', ''),
            'pincode': data.get('pincode', ''),
            'dateOfBirth': data.get('dateOfBirth', ''),
            'purchases': data.get('purchases', 0),
            'totalPurchases': data.get('totalPurchases', 0),
            'lastPurchase': data.get('lastPurchase', datetime.utcnow().date().isoformat()),
            'posBalance': data.get('posBalance', 0)
        }
        result = customers_collection.insert_one(customer_data)
        customer_data['_id'] = str(result.inserted_id)
        return jsonify({"message": "Customer added successfully", "customer": customer_data}), 201
    except Exception as e:
        logger.error(f"Error in add_customer: {e}")
        return jsonify({"error": "Failed to add customer"}), 500
@app.route('/api/customers/<id>', methods=['PUT'])
def update_customer(id):
    try:
        data = request.json
        update_data = {
            'name': data.get('name', ''),
            'phone': data.get('phone', ''),
            'email': data.get('email', ''),
            'address': data.get('address', ''),
            'city': data.get('city', ''),
            'pincode': data.get('pincode', ''),
            'dateOfBirth': data.get('dateOfBirth', ''),
            'purchases': data.get('purchases', 0),
            'totalPurchases': data.get('totalPurchases', 0),
            'lastPurchase': data.get('lastPurchase', ''),
            'posBalance': data.get('posBalance', 0)
        }
        if not all([update_data['name'], update_data['phone']]):
            return jsonify({"error": "Name and phone number are required"}), 400
        result = customers_collection.update_one({"_id": ObjectId(id)}, {"$set": update_data})
        if result.matched_count == 0:
            return jsonify({"error": "Customer not found"}), 404
        return jsonify({"message": "Customer updated successfully"})
    except Exception as e:
        logger.error(f"Error in update_customer: {e}")
        return jsonify({"error": "Failed to update customer"}), 500
@app.route('/api/customers/<id>', methods=['DELETE'])
def delete_customer(id):
    try:
        result = customers_collection.delete_one({"_id": ObjectId(id)})
        if result.deleted_count == 0:
            return jsonify({"error": "Customer not found"}), 404
        return jsonify({"message": "Customer deleted successfully"})
    except Exception as e:
        logger.error(f"Error in delete_customer: {e}")
        return jsonify({"error": "Failed to delete customer"}), 500
@app.route('/api/customers/search', methods=['GET'])
def search_customers():
    try:
        query = request.args.get('query', '')
        customers = list(customers_collection.find({
            '$or': [
                {'name': {'$regex': query, '$options': 'i'}},
                {'phone': {'$regex': query, '$options': 'i'}}
            ]
        }))
        for customer in customers:
            customer['_id'] = str(customer['_id'])
        return jsonify(customers)
    except Exception as e:
        logger.error(f"Error in search_customers: {e}")
        return jsonify({"error": "Failed to search customers"}), 500
# Import Products Route
@app.route('/api/import/products', methods=['POST'])
def import_products_from_excel():
    if 'file' not in request.files:
        return jsonify({"error": "No Excel file provided"}), 400
    excel_file = request.files['file']
    if excel_file.filename == '':
        return jsonify({"error": "No selected Excel file"}), 400
    if not allowed_excel_file(excel_file.filename):
        return jsonify({"error": "Invalid file type. Please upload an Excel file (.xlsx or .xls)"}), 400
    try:
        file_stream = io.BytesIO(excel_file.read())
        wb = openpyxl.load_workbook(file_stream)
        sheet = wb.active
        headers = [cell.value for cell in sheet[1]]
        imported_count = 0
        errors = []
        expected_headers_map = {
            "name": "name",
            "price (₹)": "price",
            "stock quantity": "stock",
            "category": "category",
            "supplier": "supplier",
            "minimum stock level": "minStock",
            "barcode": "barcode",
            "model": "model",
            "type": "type",
            "accessory type": "accessoryType",
            "image path": "image_path"
        }
        header_to_internal_key = {h.lower().strip(): expected_headers_map.get(h.lower().strip()) for h in headers if h}
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            row_values = [cell.value for cell in row]
            product_data = {}
            for header_key_excel, internal_key in header_to_internal_key.items():
                if internal_key:
                    try:
                        col_index = [h.lower().strip() for h in headers].index(header_key_excel)
                        if col_index < len(row_values):
                            product_data[internal_key] = row_values[col_index]
                    except ValueError:
                        continue
            try:
                product_data['name'] = str(product_data.get('name', '')).strip()
                if not product_data['name']:
                    errors.append(f"Row {row_idx}: Product Name is required")
                    continue
                product_data['price'] = float(product_data.get('price', 0) or 0)
                product_data['stock'] = int(product_data.get('stock', 0) or 0)
                product_data['category'] = str(product_data.get('category', '')).strip()
                product_data['supplier'] = str(product_data.get('supplier', '')).strip()
                product_data['minStock'] = int(product_data.get('minStock', 5) or 5)
                product_data['barcode'] = str(product_data.get('barcode', f"BC{int(time.time())}")).strip()
                image_path = str(product_data.get('image_path', '')).strip()
                if image_path:
                    full_path = os.path.join(app.config['UPLOAD_FOLDER'], image_path)
                    if os.path.exists(full_path):
                        product_data['image_path'] = image_path
                        logger.info(f"Image path valid for import: {full_path}")
                    else:
                        errors.append(f"Row {row_idx}: Image file '{image_path}' not found in Uploads folder")
                        product_data['image_path'] = None
                        logger.warning(f"Image not found for import: {full_path}")
                else:
                    product_data['image_path'] = None
                if product_data['category'].lower() == "mobile":
                    model_val = str(product_data.get('model', '')).strip()
                    type_val = str(product_data.get('type', '')).strip()
                    if model_val:
                        product_data['model'] = model_val
                        if not mobiles_collection.find_one({"name": model_val}):
                            mobiles_collection.insert_one({"name": model_val, "category": "Mobile"})
                    else:
                        errors.append(f"Row {row_idx}: Mobile product requires a 'Model'")
                        continue
                    product_data['type'] = type_val
                    product_data['accessoryType'] = ''
                elif product_data['category'].lower() == "accessories":
                    accessory_type_val = str(product_data.get('accessoryType', '')).strip()
                    type_val = str(product_data.get('type', '')).strip()
                    if accessory_type_val:
                        product_data['accessoryType'] = accessory_type_val
                        acc_model, acc_name = ("", accessory_type_val)
                        if ' - ' in accessory_type_val:
                            parts = accessory_type_val.split(' - ', 1)
                            if len(parts) == 2:
                                acc_model, acc_name = parts
                        if not accessories_collection.find_one({"accessoryName": acc_name, "accessoryModel": acc_model}):
                            accessories_collection.insert_one({
                                "accessoryName": acc_name,
                                "accessoryModel": acc_model,
                                "category": "Accessories",
                                "type": type_val
                            })
                    else:
                        errors.append(f"Row {row_idx}: Accessory product requires an 'Accessory Type'")
                        continue
                    product_data['model'] = ''
                    product_data['type'] = type_val
                else:
                    product_data['model'] = ''
                    product_data['accessoryType'] = ''
                    product_data['type'] = ''
                result = products_collection.insert_one(product_data)
                if product_data['stock'] > 0:
                    stock_additions_collection.insert_one({
                        "product_id": str(result.inserted_id),
                        "quantity": product_data['stock'],
                        "date": datetime.now(timezone.utc).isoformat()
                    })
                imported_count += 1
            except ValueError as ve:
                errors.append(f"Row {row_idx}: Data conversion error - {ve}")
            except Exception as e:
                errors.append(f"Row {row_idx}: Unexpected error - {e}")
        message = f"Successfully imported {imported_count} products"
        if errors:
            message += f". Errors: {'; '.join(errors)}"
            logger.error(f"Excel import errors: {errors}")
            return jsonify({"message": message, "errors": errors}), 200
        return jsonify({"message": message}), 200
    except Exception as e:
        logger.error(f"Error processing Excel file: {e}")
        return jsonify({"error": f"Failed to process Excel file: {e}"}), 500
# Report Routes
def get_month_range(year, month):
    start = datetime(year, month, 1)
    if month == 12:
        end = datetime(year + 1, 1, 1) - timedelta(days=1)
    else:
        end = datetime(year, month + 1, 1) - timedelta(days=1)
    return start, end
def get_year_range(year):
    start = datetime(year, 1, 1)
    end = datetime(year, 12, 31)
    return start, end
@app.route('/api/reports/monthly', methods=['GET'])
def get_monthly_report():
    try:
        year = int(request.args.get('year'))
        month = int(request.args.get('month'))
        start, end = get_month_range(year, month)
  
        sales_pipeline = [
            {"$match": {"timestamp": {"$gte": start.isoformat(), "$lte": end.isoformat()}}},
            {"$unwind": "$items"},
            {"$group": {
                "_id": "$items.id",
                "totalSold": {"$sum": "$items.quantity"},
                "totalRevenue": {"$sum": {"$multiply": ["$items.price", "$items.quantity"]}}
            }}
        ]
        sales_data = list(sales_collection.aggregate(sales_pipeline))
        for sale in sales_data:
            sale['_id'] = str(sale['_id'])
  
        additions_pipeline = [
            {"$match": {"date": {"$gte": start.isoformat(), "$lte": end.isoformat()}}},
            {"$group": {"_id": "$product_id", "totalAdded": {"$sum": "$quantity"}}}
        ]
        additions_data = list(stock_additions_collection.aggregate(additions_pipeline))
        for addition in additions_data:
            addition['_id'] = str(addition['_id'])
  
        current_stock = list(products_collection.find({}, {"_id": 1, "stock": 1, "name": 1}))
        for stock in current_stock:
            stock['_id'] = str(stock['_id'])
  
        report = {
            "sales": sales_data,
            "additions": additions_data,
            "current_stock": current_stock,
            "period": f"{start.strftime('%B %Y')}"
        }
        return jsonify(report)
    except Exception as e:
        logger.error(f"Error generating monthly report: {e}")
        return jsonify({"error": str(e)}), 500
@app.route('/api/reports/yearly', methods=['GET'])
def get_yearly_report():
    try:
        year = int(request.args.get('year'))
        start, end = get_year_range(year)
  
        sales_pipeline = [
            {"$match": {"timestamp": {"$gte": start.isoformat(), "$lte": end.isoformat()}}},
            {"$unwind": "$items"},
            {"$group": {
                "_id": "$items.id",
                "totalSold": {"$sum": "$items.quantity"},
                "totalRevenue": {"$sum": {"$multiply": ["$items.price", "$items.quantity"]}}
            }}
        ]
        sales_data = list(sales_collection.aggregate(sales_pipeline))
        for sale in sales_data:
            sale['_id'] = str(sale['_id'])
  
        additions_pipeline = [
            {"$match": {"date": {"$gte": start.isoformat(), "$lte": end.isoformat()}}},
            {"$group": {"_id": "$product_id", "totalAdded": {"$sum": "$quantity"}}}
        ]
        additions_data = list(stock_additions_collection.aggregate(additions_pipeline))
        for addition in additions_data:
            addition['_id'] = str(addition['_id'])
  
        current_stock = list(products_collection.find({}, {"_id": 1, "stock": 1, "name": 1}))
        for stock in current_stock:
            stock['_id'] = str(stock['_id'])
   
        report = {
            "sales": sales_data,
            "additions": additions_data,
            "current_stock": current_stock,
            "period": f"{year}"
        }
        return jsonify(report)
    except Exception as e:
        logger.error(f"Error generating yearly report: {e}")
        return jsonify({"error": str(e)}), 500
@app.route('/api/email', methods=['GET'])
def get_email_settings():
    try:
        email_settings = email_collection.find_one()
        if email_settings:
            return jsonify({
                'emailAddress': email_settings.get('emailAddress', ''),
                'fromEmailAddress': email_settings.get('fromEmailAddress', ''),
                'appPassword': email_settings.get('appPassword', '')
            })
        return jsonify({
                'emailAddress': '',
                'fromEmailAddress': '',
                'appPassword': ''
            })
    except Exception as e:
        logger.error(f"Error fetching email settings: {e}")
        return jsonify({'error': 'Failed to fetch email settings'}), 500
@app.route('/api/email', methods=['POST'])
def save_email_settings():
    try:
        data = request.get_json()
        email_address = data.get('emailAddress', '')
        from_email_address = data.get('fromEmailAddress', '')
        app_password = data.get('appPassword', '')
        email_collection.update_one(
            {},
            {'$set': {
                'emailAddress': email_address,
                'fromEmailAddress': from_email_address,
                'appPassword': app_password
            }},
            upsert=True
        )
        return jsonify({
            'emailAddress': email_address,
            'fromEmailAddress': from_email_address,
            'appPassword': app_password
        })
    except Exception as e:
        logger.error(f"Error saving email settings: {e}")
        return jsonify({'error': 'Failed to save email settings'}), 500
@app.route('/api/email', methods=['DELETE'])
def delete_email_settings():
    try:
        result = email_collection.delete_many({})
        return jsonify({"message": "Email settings deleted successfully", "deletedCount": result.deleted_count})
    except Exception as e:
        logger.error(f"Error deleting email settings: {e}")
        return jsonify({"error": "Failed to delete email settings"}), 500
def serialize_doc(doc):
    serialized = {}
    for key, value in doc.items():
        if isinstance(value, ObjectId):
            serialized[key] = str(value)
        elif isinstance(value, datetime):
            serialized[key] = value.isoformat()
        elif isinstance(value, dict):
            serialized[key] = serialize_doc(value)
        elif isinstance(value, list):
            serialized[key] = [serialize_doc(item) if isinstance(item, dict) else item for item in value]
        else:
            serialized[key] = value
    return serialized
def create_backup_file():
    collections = {
        'products': list(products_collection.find()),
        'mobiles': list(mobiles_collection.find()),
        'accessories': list(accessories_collection.find()),
        'sales': list(sales_collection.find()),
        'customers': list(customers_collection.find()),
        'settings': list(settings_collection.find()),
        'print': list(print_collection.find()),
        'email': list(email_collection.find()),
        'stock_additions': list(stock_additions_collection.find()),
    }
    serialized_collections = {}
    for collection_name, docs in collections.items():
        serialized_docs = [serialize_doc(doc) for doc in docs]
        serialized_collections[collection_name] = serialized_docs
        logger.info(f"Serialized {len(serialized_docs)} documents for collection {collection_name}")
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'backup_{timestamp}.xlsx'
    filepath = os.path.join(BACKUP_DIR, filename)
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        for collection_name, docs in serialized_collections.items():
            if docs: # Only write non-empty collections
                logger.info(f"Writing {len(docs)} documents to sheet {collection_name}")
                df = pd.DataFrame(docs)
                df.to_excel(writer, sheet_name=collection_name[:31], index=False) # Sheet names limited to 31 chars
            else:
                logger.info(f"Skipping empty collection {collection_name}")
    return filepath
def send_backup_email(filepath):
    email_settings = email_collection.find_one()
    if not email_settings:
        logger.error("No email settings found")
        return False
    from_email = email_settings.get('fromEmailAddress')
    app_password = email_settings.get('appPassword')
    backup_settings = settings_collection.find_one({"key": "backup_settings"})
    to_emails = []
    if backup_settings and backup_settings.get('backupEmail'):
        to_emails = [email.strip() for email in backup_settings.get('backupEmail', '').split(',') if email.strip()]
    if not to_emails:
        default_to_email = email_settings.get('emailAddress', '')
        if not default_to_email:
            logger.error("No backup email or default email configured")
            return False
        to_emails = [default_to_email]
    msg = MIMEMultipart()
    msg['From'] = from_email
    msg['Subject'] = "Automated Backup"
    body = "Attached is the latest automated backup."
    msg.attach(MIMEText(body, 'plain'))
    with open(filepath, "rb") as attachment:
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(attachment.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f"attachment; filename= {os.path.basename(filepath)}")
        msg.attach(part)
    try:
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(from_email, app_password)
        text = msg.as_string()
        for to_email in to_emails:
            msg['To'] = to_email
            server.sendmail(from_email, to_email, text)
        server.quit()
        logger.info("Backup email sent successfully to all recipients")
        return True
    except Exception as e:
        logger.error(f"Failed to send backup email: {e}")
        return False
def perform_backup_and_send():
    filepath = create_backup_file()
    backup_metadata = {
        'timestamp': datetime.now(),
        'filename': os.path.basename(filepath),
        'filepath': filepath,
        'type': 'automatic'
    }
    backup_collection.insert_one(backup_metadata)
    send_backup_email(filepath)
@app.route('/api/backup', methods=['GET'])
def handle_backup():
    if not backup_collection or not all([products_collection, mobiles_collection, accessories_collection, sales_collection, customers_collection, settings_collection, print_collection, email_collection, stock_additions_collection]):
        logger.error("Database connection or collections missing")
        return jsonify({'error': 'Database connection failed or collections missing'}), 500
    try:
        filepath = create_backup_file()
        backup_metadata = {
            'timestamp': datetime.now(),
            'filename': os.path.basename(filepath),
            'filepath': filepath,
            'type': 'manual'
        }
        backup_collection.insert_one(backup_metadata)
        logger.info(f"Backup metadata saved: {os.path.basename(filepath)}")
        email_sent = send_backup_email(filepath)
        download_url = f'http://localhost:5000/static/backups/{os.path.basename(filepath)}'
        message = 'Backup created successfully'
        if email_sent:
            message += ' and sent to email'
        return jsonify({
            'message': message,
            'downloadUrl': download_url
        })
    except Exception as e:
        logger.error(f"Error creating backup: {str(e)}")
        return jsonify({'error': f'Failed to create backup: {str(e)}'}), 500
@app.route('/api/backup/settings', methods=['GET'])
def get_backup_settings():
    try:
        backup_settings = settings_collection.find_one({"key": "backup_settings"})
        if backup_settings:
            return jsonify({
                "backupEmail": backup_settings.get("backupEmail", ""),
                "backupOptions": backup_settings.get("backupOptions", {"daily": False, "weekly": False, "monthly": False}),
                "dailyInterval": backup_settings.get("dailyInterval", 24),
                "weeklyDay": backup_settings.get("weeklyDay", "mon"),
                "monthlyDay": backup_settings.get("monthlyDay", 1)
            })
        return jsonify({
            "backupEmail": "",
            "backupOptions": {"daily": False, "weekly": False, "monthly": False},
            "dailyInterval": 24,
            "weeklyDay": "mon",
            "monthlyDay": 1
        })
    except Exception as e:
        logger.error(f"Error fetching backup settings: {e}")
        return jsonify({"error": "Failed to fetch backup settings"}), 500
@app.route('/api/backup/settings', methods=['POST'])
def save_backup_settings():
    try:
        data = request.json
        backup_email = data.get('backupEmail', '')
        backup_options = data.get('backupOptions', {"daily": False, "weekly": False, "monthly": False})
        daily_interval = int(data.get('dailyInterval', 24))
        weekly_day = data.get('weeklyDay', 'mon')
        monthly_day = int(data.get('monthlyDay', 1))
        settings_collection.update_one(
            {"key": "backup_settings"},
            {"$set": {
                "backupEmail": backup_email,
                "backupOptions": backup_options,
                "dailyInterval": daily_interval,
                "weeklyDay": weekly_day,
                "monthlyDay": monthly_day
            }},
            upsert=True
        )
        # Reschedule jobs
        if scheduler.get_job('daily_backup'):
            scheduler.remove_job('daily_backup')
        if scheduler.get_job('weekly_backup'):
            scheduler.remove_job('weekly_backup')
        if scheduler.get_job('monthly_backup'):
            scheduler.remove_job('monthly_backup')
        if backup_options.get('daily', False):
            scheduler.add_job(perform_backup_and_send, 'interval', hours=daily_interval, id='daily_backup', replace_existing=True)
        if backup_options.get('weekly', False):
            scheduler.add_job(perform_backup_and_send, 'cron', day_of_week=weekly_day, hour=0, minute=0, id='weekly_backup', replace_existing=True)
        if backup_options.get('monthly', False):
            scheduler.add_job(perform_backup_and_send, 'cron', day=monthly_day, hour=0, minute=0, id='monthly_backup', replace_existing=True)
        return jsonify({"message": "Backup settings saved successfully"})
    except Exception as e:
        logger.error(f"Error saving backup settings: {e}")
        return jsonify({"error": "Failed to save backup settings"}), 500
@app.route('/api/backup/settings', methods=['DELETE'])
def delete_backup_settings():
    try:
        result = settings_collection.delete_one({"key": "backup_settings"})
        if scheduler.get_job('daily_backup'):
            scheduler.remove_job('daily_backup')
        if scheduler.get_job('weekly_backup'):
            scheduler.remove_job('weekly_backup')
        if scheduler.get_job('monthly_backup'):
            scheduler.remove_job('monthly_backup')
        return jsonify({"message": "Backup settings deleted successfully", "deletedCount": result.deleted_count})
    except Exception as e:
        logger.error(f"Error deleting backup settings: {e}")
        return jsonify({"error": "Failed to delete backup settings"}), 500
@app.route('/api/backup/last', methods=['GET'])
def get_last_backup():
    try:
        last_backup = backup_collection.find_one(sort=[('timestamp', -1)])
        if last_backup:
            return jsonify({"lastBackupDate": last_backup['timestamp'].isoformat()})
        return jsonify({"lastBackupDate": None})
    except Exception as e:
        logger.error(f"Error fetching last backup: {e}")
        return jsonify({"error": "Failed to fetch last backup"}), 500
@app.route('/static/backups/<filename>')
def serve_backup_file(filename):
    try:
        return send_from_directory(BACKUP_DIR, filename, as_attachment=True)
    except Exception as e:
        logger.error(f"Error serving backup file {filename}: {str(e)}")
        return jsonify({'error': f'Failed to serve backup file: {str(e)}'}), 404
# Scheduler setup
def start_scheduler():
    scheduler.start()
    backup_settings = settings_collection.find_one({"key": "backup_settings"})
    if backup_settings:
        backup_options = backup_settings.get("backupOptions", {})
        daily_interval = backup_settings.get("dailyInterval", 24)
        weekly_day = backup_settings.get("weeklyDay", "mon")
        monthly_day = backup_settings.get("monthlyDay", 1)
        if backup_options.get('daily', False):
            scheduler.add_job(perform_backup_and_send, 'interval', hours=daily_interval, id='daily_backup', replace_existing=True)
        if backup_options.get('weekly', False):
            scheduler.add_job(perform_backup_and_send, 'cron', day_of_week=weekly_day, hour=0, minute=0, id='weekly_backup', replace_existing=True)
        if backup_options.get('monthly', False):
            scheduler.add_job(perform_backup_and_send, 'cron', day=monthly_day, hour=0, minute=0, id='monthly_backup', replace_existing=True)
    logger.info("Scheduler started.")
# Main execution
def shutdown_server():
    global db, fs
    if db:
        db.client.close()
        db = None
        fs = None
    logger.info("Server shutting down gracefully.")
# Main execution with shutdown handler
if __name__ == '__main__':
    try:
        if db:
            start_scheduler()
        logger.info(f"Serving static files from: {app.static_folder}")
        if getattr(sys, 'frozen', False):
            logger.info("Running as frozen executable, using Waitress")
            from waitress import serve
            serve(app, host='0.0.0.0', port=5000, threads=8)
        else:
            logger.info("Running in development mode, using Flask")
            app.run(host='0.0.0.0', port=5000, debug=True)
    except KeyboardInterrupt:
        shutdown_server()
    except Exception as e:
        logger.error(f"Server error: {e}")
        shutdown_server()